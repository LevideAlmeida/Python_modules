# Doc: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active  # type: ignore

sheet_name = 'Python Sheet'

# Create sheet in 0 index
workbook.create_sheet(sheet_name, 0)

# Select workbook
worksheet: Worksheet = workbook[sheet_name]

# Remove workbook
workbook.remove(workbook['Sheet'])

worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'Age')
worksheet.cell(1, 3, 'Grade')

students = [
    # Name     # Age   # Grade
    ['Levi',      16,    10],
    ['Vitoria',   17,    9],
    ['Duda',      15,    6],
    ['Policarpo', 16,    4],
]

# for i, students_row in enumerate(students, start=2):
#     for j, students_column in enumerate(students_row, start=1):
#         worksheet.cell(i, j, students_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
