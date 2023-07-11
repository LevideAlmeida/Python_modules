# Doc: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook: Workbook = load_workbook(WORKBOOK_PATH)

sheet_name = 'Python Sheet'

# Select workbook
worksheet: Worksheet = workbook[sheet_name]


for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Poli':
            worksheet.cell(cell.row, 3, 0)  # type: ignore

    print()

worksheet['A5'].value = 'Poli'

workbook.save(WORKBOOK_PATH)
